import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side


wb = openpyxl.load_workbook('Теория.xlsx')
sheet = wb.active


theory_set = set()


for cell in sheet['A']:
    theory_set.add(cell.value)


wb = openpyxl.load_workbook('Практика.xlsx')
sheet = wb.active


fact_set = set()


for cell in sheet['A']:
    fact_set.add(cell.value)


s_lght = list(theory_set - fact_set)
v_sale = list(fact_set - theory_set)


wb = openpyxl.Workbook()
sheet = wb.active


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


al = Alignment(horizontal='center')

def right (x, y):
    sheet.cell(row=x, column=y).border = thin_border
    sheet.cell(row=x, column=y).alignment = al


sheet.cell(row=2, column=2).value = "S-Light"
right(2, 2)

sheet.cell(row=2, column=4).value = "V-Sale"
right(2, 4)



for i in range (len(s_lght)):
    sheet.cell(row=3+i, column=2).value = s_lght[i]
    right(3+i, 2)


for i in range (len(v_sale)):
    sheet.cell(row=3+i, column=4).value = v_sale[i]
    right(3 + i, 4)


wb.save('Результат.xlsx')








